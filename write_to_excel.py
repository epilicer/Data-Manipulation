
import csv
import openpyxl

from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
# from openpyxl.drawing.line import LineProperties


def csv_to_excel(csv_files, excel_file):
    """
    read list of csv files into an excel file
    https://www.blog.pythonlibrary.org/2021/09/25/converting-csv-to-excel-with-python/
    :param csv_files: list of csv files as input
    :param excel_file: output excel file
    :return: None
    """
    workbook = openpyxl.Workbook()    # create an empty book (excel)
    workbook.remove(workbook.active)  # remove empty sheet when creating an empty book
    # iterate over list of csv files
    for csv_file in csv_files:
        csv_data = []
        # read single csv file
        with open(csv_file, 'r', encoding='utf-8') as file_obj:   # encoding required to eliminate weird chars
            reader = csv.reader(file_obj)
            for row in reader:
                csv_data.append(row)
        csv_file = csv_file.split('/')[-1][:-4]   # rewrite sheet name
        worksheet = workbook.create_sheet(csv_file)   # create a separate sheet
        # write single csv data
        for row in csv_data:
            worksheet.append(row)
    # iteration finished, save the book
    workbook.save(excel_file)
    return None


def plot_scatter_chart(workbook, worksheets, graph_title, x_cells, y_cells):
    """
    plot function for scatter chart
    :param workbook:
    :param worksheets:
    :param graph_title:
    :param x_cells:
    :param y_cells:
    :return:
    """
    chart = ScatterChart()
    chart.style = 13     # for scatter chart
    chart.title = graph_title
    chart.y_axis.title = 'Value (Lin)'
    chart.legend.position = 'tr'
    chart.layout = Layout( manualLayout=ManualLayout(x=0.25, y=0.25, h=1, w=1))
    chart.height = 20    # cm
    chart.width = 35     # cm
    # some cosmeticity, if needed
    # https://www.color-hex.com/color-names.html
    # chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True), bwMode='black')
    # chart.graphical_properties = GraphicalProperties(bwMode='black')
    # chart.plot_area.graphicalProperties = GraphicalProperties(solidFill="999999")  # Gray
    # chart.x_axis.majorGridlines.spPr = GraphicalProperties(solidFill='FFFFFF')  # White
    # chart.y_axis.majorGridlines.spPr = GraphicalProperties(solidFill='FFFFFF')    # White
    # chart.x_axis.minorGridlines = ChartLines(GraphicalProperties(solidFill='0048ba') )
    # chart.y_axis.minorGridlines = ChartLines(GraphicalProperties(solidFill='0048ba') )
    # chart.varyColors = True
    # chart.x_axis.scaling.logBase = 10
    # chart.x_axis.scaling.min = 100
    # chart.x_axis.scaling.max = 20000
    worksheet = workbook[worksheets[0]]
    chart.x_axis.title = str(worksheet['A6'].value)
    for ws in worksheets:
        worksheet = workbook[ws]
        x_values = Reference(worksheet, min_col=x_cells[0], min_row=x_cells[1], max_row=x_cells[2])
        values = Reference(worksheet, min_col=y_cells[0], min_row=y_cells[1], max_row=y_cells[2])
        series = Series(values, x_values, title_from_data=True)
        # series.marker = openpyxl.chart.marker.Marker('x')   # for marker plot
        # series.graphicalProperties.line.noFill = True
        series.legend = ws
        chart.series.append(series)
    # chart.legend.LegendEntry = (LegendEntry(0, worksheets[0]),
    #                             LegendEntry(1, worksheets[1]))
    workbook.create_sheet(graph_title)
    worksheet = workbook[graph_title]
    worksheet.add_chart(chart, 'A1')
    # cs = workbook.create_chartsheet(graph_title)   # this did not work
    # cs.add_chart(chart)
    return None


if __name__ == '__main__':
    # user inputs
    csv_file_names = ['data/data1_gated.csv', 'data/data2_gated.csv']
    xls_file_name = 'books.xlsx'
    
    # call to read & write method
    # csv_to_excel(csv_file_names, xls_file_name)

    # if excel is already there, open and do plotting
    workbook = openpyxl.load_workbook(xls_file_name)
    worksheets = [c.split('/')[-1][:-4] for c in csv_file_names]
    
    graph_title = 'S222 (Log)'
    x_cells = [1, 6, 28]  # min_col, min_row, max_row
    y_cells = [3, 6, 28]  # min_col, min_row, max_row
    plot_scatter_chart(workbook, worksheets, graph_title, x_cells, y_cells)
    
    graph_title = 'S33 (Log)'
    x_cells = [1, 6, 28]  # min_col, min_row, max_row
    y_cells = [4, 6, 28]  # min_col, min_row, max_row
    plot_scatter_chart(workbook, worksheets, graph_title, x_cells, y_cells)
    
    graph_title = 'S44 (Log)'
    x_cells = [1, 6, 28]  # min_col, min_row, max_row
    y_cells = [5, 6, 28]  # min_col, min_row, max_row
    plot_scatter_chart(workbook, worksheets, graph_title, x_cells, y_cells)

    workbook.save(xls_file_name)


############################################################
# Open excel and check/remove single quotes
# Select column and -> Convert Text to Column
############################################################

